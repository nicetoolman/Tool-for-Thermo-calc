import pandas as pd

def exist_or_create_excel(file_path, columns):
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    import os
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        # 如果文件不存在，创建一个特定格式的Excel文件
        df = pd.DataFrame(columns=columns)# 替换为您的列名
    
        # 保存为Excel文件
        df.to_excel(file_path, index=False)
                
        # 加载工作簿并选择活动工作表
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 设置列标题的字体
        font_name = 'Yu Gothic'
        font_size = 11
        font = Font(name=font_name, size=font_size)
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(columns)):
            for cell in col:
                cell.font = font
                
        # 保存工作簿
        wb.save(file_path)
        
        print(f'文件 {file_path} 已创建')
        
        import sys
        sys.exit()
    else:
        print(f'文件 {file_path} 已存在')

def find_elements_from_filename(filename):    
    filename = filename.split()[0].split('-')
    element1 = filename[0]
    element2 = filename[1]    
    return element1, element2

def find_common_elements(list1, list2):
    common_elements = [element for element in list1 if element in list2]
    unique_to_list1 = [element for element in list1 if element not in list2]
    unique_to_list2 = [element for element in list2 if element not in list1]
    
    if common_elements:
        return common_elements, unique_to_list1, unique_to_list2
    else:
        return []

def find_phases(transition) -> dict:
    '''
    只针对固液共存区的实验数据
    这个函数的作用是把相转换的式子拆开并把不重复的相读入一个字典型中
    例如：
    print(find_phases("LIQUID->LIQUID+DIS_FCC"))
    输出为
    {'phase1': 'LIQUID', 'phase2': 'DIS_FCC'}
    '''
    try:
        transition = transition.strip()
        transition = transition.replace(' ', '')
        transition = transition.split('->')
        pre = transition[0]
        post = transition[1]
        pre = pre.split('+')
        post = post.split('+')
    except:
        print('transition formula goes wrong, phases can only be connected by "->" and "+"')
        import sys
        sys.exit()
    '''
    上面的代码的作用是：
    - 一定程度上规范化转换式，去掉字符串前后和之中的空格
    - 拆开表达式，只保留相的部分
    '''

    try:
        if len(pre) == 1:    
            phase_dict = dict.fromkeys(pre + post)
        elif len(pre) == 2:            
            if pre[0] == 'LIQUID':
                phase_dict = dict.fromkeys(pre + post)           
            elif pre[1] == 'LIQUID':
                pre[0], pre[1] = pre[1], pre[0]
                phase_dict = dict.fromkeys(pre + post)             
    except:
        print('transition formula goes wrong, check the number of phases')
        import sys
        sys.exit()                  
    # print(list(phase_dict))
    '''
    上面的代码的目的是不重复的，按一定顺序地把相存进一个字典
    - 这里用了dict.fromkeys，所以读入的相名会被当做key来读入字典型，去掉重复的并保持原来的顺序
    - 这里主要考虑的是固液共存区的实验数据
    - 会保持LIQUID在第一个
    '''

    try:
        phases = {}
        i = 0
        for phase in list(phase_dict):
            phases[f'phase{i+1}'] = phase
            i = i + 1
    except:
        print('transition formula goes wrong')
        import sys
        sys.exit()
    # print(phases[f"phase{i+1}"])
    return phases, pre, post
    '''
    这里的作用就是循环读入phase的名字到的（key）phase1，phase2，phase3...中
    '''

def convert_to_pop(filename, pop_file):
        """读取 Excel 文件并返回 DataFrame"""
        df = pd.read_excel(filename)
        """要确定一个excel文件的命名规则，即：
        两个元素要放在最前面，用'-'连接
        而且元素排列顺序与对应元素组成顺序一致
        比如： W-Co experimental data.xlsx
        文件内顺序为x(W), x(Co)"""
        #主要目的是找到最前面的两个元素
        #所以要先按默认的空格分开，取出最前面的Co-W部分
        #再把-两边的元素分别取出来
        filename = filename.split()[0].split('-')
        element1 = filename[0]
        element2 = filename[1]
        elements = dict()
        x_element1 = f"x_element_{element1}"#动态读入不同二元系的元素
        x_element2 = f"x_element_{element2}"#同上
        
        with open(pop_file, 'w') as file: #用with的好处就是操作结束自动关闭文件，而这里的操作类型是w，也就是写入，在没有该文件夹的情况下会创建一个
            # 写入 POP 文件头部信息（根据实际需要修改）
            # 遍历 DataFrame 的每一行，写入 POP 文件
            HEADER = input('please type in popfile info: ')
            file.write(HEADER + "\n\n")#该pop文件的注释信息
            file.write('ENTER_SYMBOL CONSTANT DX=0.02, P0=101325, DH=500, DT=10\n\n')#定义一些常数
            i = 1#循环读取DataFrame（pandas的数据类型）的每一行，设置一个循环变量来记录循环次数
            for index, row in df.iterrows():#遍历每一行
                if row.isnull().all():#检查每一行是否是空白行，如果是则结束读取。这个地方可以用，但是跟我想的不一样，有一些微妙的问题
                    print(f"Row {index + 1} is empty. Exiting the program.")                    
                else:
                    '''
                    下面这一段都是在从excel把每列的数据取出来
                    '''                                 
                    elements[x_element1] = row['x(' + element1 +')']
                #    print(row['x(' + element1 +')'])
                    elements[x_element2] = row['x(' + element2 +')']    
                #    print(row['x(' + element2 +')'])
                #    print(elements.items())
                    temperature_C = row['Temperature/ºC']
                    temperature_K = row['Temperature/K']
                    lable = row['Lable']
                    transition = row['Transition']
                    find_phases_result = find_phases(transition)
                    phases = find_phases_result[0]
                    pre = find_phases_result[1]
                    post = find_phases_result[2]
                    #print(phases, pre, post)
                    phase1 = phases['phase1']
                    phase2 = phases['phase2']
                    if len(phases) == 3:
                        phase3 = phases['phase3']
                    paper = row['Paper']
                    reporter = row['Reporter']
                    methodology = row['Methodology']
                    '''
                    取完了
                    '''
                    #print(phase1, 
                    #      phase2, 
                    #      phase3 if 'phase3' in locals() else "")
                       
                    file.write(f"CREATE_NEW_EQUILIBRIUM, {i}, 1\n")#pop文件的计数    
                    
                    if  find_common_elements(pre, post):#根据相变反应的前后的反应物和生成物的种类来决定如何写change condition
                        #如果反应前后有相同的相，那么该项fix1， 则另外的相设定fix0
                        common_elements, unique_to_list1, unique_to_list2 = find_common_elements(pre, post)
                        #print(common_elements, unique_to_list1, unique_to_list2)
                        file.write(f"CHANGE_STATUS PHASE {common_elements[0]} = FIX 1\n")
                        if unique_to_list1 == [] and unique_to_list2:
                            file.write(f"CHANGE_STATUS PHASE {unique_to_list2[0]} = FIX 0\n")
                        elif unique_to_list1 and unique_to_list2 == []:     
                            file.write(f"CHANGE_STATUS PHASE {unique_to_list1[0]} = FIX 0\n")
                        elif unique_to_list1 and unique_to_list2:
                            file.write(f"CHANGE_STATUS PHASE {unique_to_list1[0]} = FIX 1\n")
                            file.write(f"CHANGE_STATUS PHASE {unique_to_list2[0]} = FIX 1\n")
                        if len(phases) == 3:
                            file.write(f"CHANGE_STATUS PHASE {unique_to_list2[0]} = FIX 1\n") 
                    else:
                        file.write(f"CHANGE_STATUS PHASE {phase1} = FIX 1\n")
                        file.write(f"CHANGE_STATUS PHASE {phase2} = FIX 1\n")
                        file.write(f"CHANGE_STATUS PHASE {phase3} = FIX 1\n")                 
                    if len(phases) == 3:#根据自由度添加set condition
                        file.write("SET_CONDITION P=P0\n")
                    if len(phases) == 2:#同上
                        file.write("SET_CONDITION P=P0\n")
                        file.write(f"SET_CONDITION x({element1})={elements[x_element1]}\n")
                    file.write(f"EXPERIMENT T={temperature_K}:DT\n")
                    file.write(f"LABEL {lable}\n")
                    file.write(f"SET_START_VALUE T={temperature_K}\n")
                    file.write("\n")  # 空行分隔每个记录
                    i = i + 1
                    
            file.write("SAVE_WORKSPACES\n")  # 添加 SAVE_WORKSPACES
            print('Transiton success!') 
            print('Nice bro, you life was just saved again from simple repetitive tasks')                 
            return 

#按规则输入文件名
datasetfile = input('please input your dataset file: ')
if len(datasetfile) == 0:
    datasetfile = "W-Co experimental data.xlsx"
#找到文件名中的元素
element1, element2 = find_elements_from_filename(datasetfile)
#定义excel的样式，为生成特定格式的Excel文件做准备
columns = ['ID', 
           f'x({element1})', 
           f'x({element2})', 
           'Temperature/ºC', 
           'Temperature/K', 
           'Lable', 
           'Transition', 
           'Phase1', 
           'Phase2', 
           'Phase3', 
           'Paper', 
           'Reporter', 
           'Methodology']  # 替换为您的列名  
#检查目录中是否有该excel文件，没有则创建一个特定格式的excel文件
exist_or_create_excel(datasetfile, columns)   
#设定输出的pop文件的文件名
popfile = f'{element1}_{element2}_popfile.pop'     
#调用转化pop文件的函数    
convert_to_pop(datasetfile, popfile)


    

