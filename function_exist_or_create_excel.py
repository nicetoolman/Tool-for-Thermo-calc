import pandas as pd

def exist_or_create_excel(file_path, columns):
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    import os
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        # 如果文件不存在，创建一个特定格式的Excel文件
        df = pd.DataFrame(columns=columns)# 替换为您的列名
    
        # 保存为Excel文件
        df.to_excel(file_path, index=False)
                
        # 加载工作簿并选择活动工作表
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 设置列标题的字体
        font_name = 'Yu Gothic'
        font_size = 11
        font = Font(name=font_name, size=font_size)
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(columns)):
            for cell in col:
                cell.font = font
                
        # 保存工作簿
        wb.save(file_path)
        
        print(f'文件 {file_path} 已创建')
    else:
        print(f'文件 {file_path} 已存在')


file_path = input('Please input the dataset file name: ')
if len(file_path) == 0:
    file_path = 'new_dataset.xlsx'
columns = ['ID', 
           'x(W)', 
           'x(Co)', 
           'Temperature/ºC', 
           'Temperature/K', 
           'Lable', 
           'Transition', 
           'Phase1', 
           'Phase2', 
           'Phase3', 
           'Paper', 
           'Reporter', 
           'Methodology']  # 替换为您的列名

exist_or_create_excel(file_path, columns)